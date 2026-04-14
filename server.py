import json
import os
import uuid
import datetime
import smtplib
from email.message import EmailMessage
from pathlib import Path

from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook, load_workbook
from supabase import create_client

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

BASE_DIR = Path(__file__).resolve().parent
MASTER_FILE = BASE_DIR / "master.xlsx"
COLLECTIONS_FILE = BASE_DIR / "collections.json"
EMAIL_CONFIG_FILE = BASE_DIR / "email_config.json"
SERVER_CONFIG_FILE = BASE_DIR / "server_config.json"
REPORTS_DIR = BASE_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True)
USER_FILE_TEMPLATE = BASE_DIR / "planilha_{equipe}.xlsx"
HEADERS = ["Data", "Equipe", "Medidor", "Tipo", "Status", "Observações", "Operador", "Coleção"]
DEFAULT_EMAIL_CONFIG = {
    "smtp_server": "",
    "smtp_port": 587,
    "smtp_user": "",
    "smtp_password": "",
    "from_address": "",
    "use_tls": True,
}
DEFAULT_SERVER_CONFIG = {
    "api_key": "",
}

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").strip()
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "").strip()
SUPABASE_TABLE = os.environ.get("SUPABASE_TABLE", "medicoes").strip()
SUPABASE_CLIENT = None
if SUPABASE_URL and SUPABASE_KEY:
    SUPABASE_CLIENT = create_client(SUPABASE_URL, SUPABASE_KEY)

app = Flask(__name__)
app.secret_key = "change-this-secret"
app.config["UPLOAD_FOLDER"] = str(BASE_DIR)


def ensure_workbook(path: Path) -> Workbook:
    if path.exists():
        wb = load_workbook(path)
        if "Dados" not in wb.sheetnames:
            ws = wb.create_sheet("Dados")
            ws.append(HEADERS)
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.append(HEADERS)
    wb.save(path)
    return wb


def append_row(path: Path, row: list[str]) -> None:
    wb = ensure_workbook(path)
    ws = wb["Dados"]
    ws.append(row)
    wb.save(path)


def load_email_config() -> dict:
    if EMAIL_CONFIG_FILE.exists():
        with open(EMAIL_CONFIG_FILE, "r", encoding="utf-8") as handle:
            return json.load(handle)
    return DEFAULT_EMAIL_CONFIG.copy()


def save_email_config(config: dict) -> None:
    with open(EMAIL_CONFIG_FILE, "w", encoding="utf-8") as handle:
        json.dump(config, handle, indent=2, ensure_ascii=False)


def ensure_email_config() -> dict:
    config = load_email_config()
    if not EMAIL_CONFIG_FILE.exists():
        save_email_config(config)
    return config


def is_email_configured(config: dict) -> bool:
    return bool(config.get("smtp_server") and config.get("smtp_user") and config.get("from_address"))


def load_server_config() -> dict:
    if SERVER_CONFIG_FILE.exists():
        with open(SERVER_CONFIG_FILE, "r", encoding="utf-8") as handle:
            return json.load(handle)
    return DEFAULT_SERVER_CONFIG.copy()


def save_server_config(config: dict) -> None:
    with open(SERVER_CONFIG_FILE, "w", encoding="utf-8") as handle:
        json.dump(config, handle, indent=2, ensure_ascii=False)


def ensure_server_config() -> dict:
    config = load_server_config()
    if not SERVER_CONFIG_FILE.exists():
        save_server_config(config)
    return config


def api_key_required() -> bool:
    config = load_server_config()
    if not config.get("api_key"):
        return False

    key = request.headers.get("X-API-KEY") or request.args.get("api_key", "")
    return key == config["api_key"]


def is_supabase_enabled() -> bool:
    return SUPABASE_CLIENT is not None


def store_supabase_entry(entry: dict) -> None:
    if not is_supabase_enabled():
        return

    payload = {
        "collection_id": entry["collection_id"],
        "equipe": entry["equipe"],
        "operador": entry["operador"],
        "medidor": entry["medidor"],
        "tipo": entry["tipo"],
        "status": entry["status"],
        "observacoes": entry["observacoes"],
        "collection_start": entry["collection_start"],
        "recorded_at": entry["recorded_at"],
    }
    response = SUPABASE_CLIENT.table(SUPABASE_TABLE).insert(payload).execute()
    if getattr(response, "error", None):
        raise RuntimeError(response.error.message or str(response.error))


def send_email_report(to_address: str, subject: str, body: str, attachment_path: Path) -> None:
    config = load_email_config()
    if not is_email_configured(config):
        raise ValueError("Configuração de email não está completa.")

    message = EmailMessage()
    message["From"] = config["from_address"]
    message["To"] = to_address
    message["Subject"] = subject
    message.set_content(body)

    with open(attachment_path, "rb") as f:
        data = f.read()

    if attachment_path.suffix.lower() == ".pdf":
        maintype, subtype = "application", "pdf"
    else:
        maintype, subtype = "text", "plain"

    message.add_attachment(data, maintype=maintype, subtype=subtype, filename=attachment_path.name)

    smtp_port = int(config.get("smtp_port", 587))
    with smtplib.SMTP(config["smtp_server"], smtp_port) as server:
        if config.get("use_tls", True):
            server.starttls()
        server.login(config["smtp_user"], config["smtp_password"])
        server.send_message(message)


def slugify(value: str) -> str:
    return "".join(c if c.isalnum() else "_" for c in value).strip("_")


def load_collections() -> dict:
    if COLLECTIONS_FILE.exists():
        with open(COLLECTIONS_FILE, "r", encoding="utf-8") as handle:
            return json.load(handle)
    return {}


def save_collections(data: dict) -> None:
    with open(COLLECTIONS_FILE, "w", encoding="utf-8") as handle:
        json.dump(data, handle, indent=2, ensure_ascii=False)


def create_pdf_report(collection: dict) -> Path:
    report_path = REPORTS_DIR / f"relatorio_{collection['id']}.pdf"
    if REPORTLAB_AVAILABLE:
        c = canvas.Canvas(str(report_path), pagesize=A4)
        width, height = A4
        y = height - 40
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, "Relatório de Coleta de Medidores")
        c.setFont("Helvetica", 11)
        y -= 30
        lines = [
            f"Equipe: {collection['equipe']}",
            f"Operador: {collection['operador']}",
            f"Início: {collection['start']}",
            f"Término: {collection.get('end', 'em andamento')}",
            f"Total de registros: {len(collection['items'])}",
            "",
        ]
        for line in lines:
            c.drawString(40, y, line)
            y -= 18

        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, y, "Data / Hora")
        c.drawString(160, y, "Medidor")
        c.drawString(300, y, "Tipo")
        c.drawString(380, y, "Status")
        c.drawString(460, y, "Observações")
        y -= 18
        c.setFont("Helvetica", 9)

        for item in collection["items"]:
            if y < 70:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica-Bold", 10)
                c.drawString(40, y, "Data / Hora")
                c.drawString(160, y, "Medidor")
                c.drawString(300, y, "Tipo")
                c.drawString(380, y, "Status")
                c.drawString(460, y, "Observações")
                y -= 18
                c.setFont("Helvetica", 9)

            c.drawString(40, y, item["timestamp"])
            c.drawString(160, y, item["medidor"][:16])
            c.drawString(300, y, item["tipo"])
            c.drawString(380, y, item["status"])
            c.drawString(460, y, item["observacoes"][:30])
            y -= 16

        c.save()
    else:
        report_path = REPORTS_DIR / f"relatorio_{collection['id']}.txt"
        with open(report_path, "w", encoding="utf-8") as handle:
            handle.write("Relatório de Coleta de Medidores\n")
            handle.write(f"Equipe: {collection['equipe']}\n")
            handle.write(f"Operador: {collection['operador']}\n")
            handle.write(f"Início: {collection['start']}\n")
            handle.write(f"Término: {collection.get('end', 'em andamento')}\n")
            handle.write(f"Total de registros: {len(collection['items'])}\n\n")
            for item in collection["items"]:
                handle.write(
                    f"{item['timestamp']} | {item['medidor']} | {item['tipo']} | {item['status']} | {item['observacoes']}\n"
                )
    return report_path


def create_entry(collection: dict, medidor: str, tipo: str, status: str, observacoes: str) -> None:
    now = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
    item = {
        "timestamp": now,
        "medidor": medidor,
        "tipo": tipo,
        "status": status,
        "observacoes": observacoes,
    }
    collection["items"].append(item)
    collection["updated"] = now

    row = [
        now,
        collection["equipe"],
        medidor,
        tipo,
        status,
        observacoes,
        collection["operador"],
        collection["id"],
    ]

    append_row(MASTER_FILE, row)
    user_file = Path(USER_FILE_TEMPLATE.format(equipe=slugify(collection["equipe"])))
    append_row(user_file, row)

    if is_supabase_enabled():
        store_supabase_entry({
            "collection_id": collection["id"],
            "equipe": collection["equipe"],
            "operador": collection["operador"],
            "medidor": medidor,
            "tipo": tipo,
            "status": status,
            "observacoes": observacoes,
            "collection_start": collection["start"],
            "recorded_at": now,
        })


@app.route("/")
def index():
    collections = load_collections()
    active = [c for c in collections.values() if c.get("status") != "finalizada"]
    return render_template("index.html", active=active)


@app.route("/settings", methods=["GET", "POST"])
def settings():
    email_config = load_email_config()
    server_config = load_server_config()
    if request.method == "POST":
        email_config["smtp_server"] = request.form.get("smtp_server", "").strip()
        email_config["smtp_port"] = int(request.form.get("smtp_port", 587))
        email_config["smtp_user"] = request.form.get("smtp_user", "").strip()
        email_config["smtp_password"] = request.form.get("smtp_password", "").strip()
        email_config["from_address"] = request.form.get("from_address", "").strip()
        email_config["use_tls"] = request.form.get("use_tls") == "on"
        server_config["api_key"] = request.form.get("api_key", "").strip()
        save_email_config(email_config)
        save_server_config(server_config)
        flash("Configurações salvas.", "success")
        return redirect(url_for("settings"))
    return render_template("settings.html", email_config=email_config, server_config=server_config)


def require_api_auth():
    config = load_server_config()
    if config.get("api_key") and not api_key_required():
        return jsonify({"error": "API key required"}), 401
    return None


def get_collection_or_404(collection_id: str):
    collections = load_collections()
    collection = collections.get(collection_id)
    if not collection:
        return jsonify({"error": "Coleta não encontrada"}), 404
    return collection


@app.route("/mobile")
def mobile():
    return render_template("mobile.html")


@app.route("/api/start", methods=["POST"])
def api_start():
    auth = require_api_auth()
    if auth:
        return auth

    data = request.get_json(silent=True) or {}
    equipe = data.get("equipe", "").strip()
    operador = data.get("operador", "").strip()
    if not equipe or not operador:
        return jsonify({"error": "Informe equipe e operador."}), 400

    collection_id = str(uuid.uuid4())
    now = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
    collection = {
        "id": collection_id,
        "equipe": equipe,
        "operador": operador,
        "start": now,
        "updated": now,
        "items": [],
        "status": "aberta",
    }
    collections = load_collections()
    collections[collection_id] = collection
    save_collections(collections)
    return jsonify({"collection_id": collection_id, "start": now}), 201


@app.route("/api/collect/<collection_id>", methods=["POST"])
def api_collect(collection_id):
    auth = require_api_auth()
    if auth:
        return auth

    collection = get_collection_or_404(collection_id)
    if isinstance(collection, tuple):
        return collection

    data = request.get_json(silent=True) or {}
    medidor = data.get("medidor", "").strip()
    tipo = data.get("tipo", "").strip()
    status = data.get("status", "").strip()
    observacoes = data.get("observacoes", "").strip()
    if not medidor or not tipo or not status:
        return jsonify({"error": "Preencha medidor, tipo e status."}), 400

    try:
        create_entry(collection, medidor, tipo, status, observacoes)
        collections = load_collections()
        collections[collection_id] = collection
        save_collections(collections)
        return jsonify({"message": "Registro salvo."}), 201
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/api/finalize/<collection_id>", methods=["POST"])
def api_finalize(collection_id):
    auth = require_api_auth()
    if auth:
        return auth

    collection = get_collection_or_404(collection_id)
    if isinstance(collection, tuple):
        return collection

    if collection.get("status") == "finalizada":
        return jsonify({"message": "Coleta já finalizada."}), 200

    collection["end"] = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
    collection["status"] = "finalizada"
    collections = load_collections()
    collections[collection_id] = collection
    save_collections(collections)
    report_path = create_pdf_report(collection)
    return jsonify({"message": "Coleta finalizada.", "report": report_path.name}), 200


@app.route("/send_email/<collection_id>", methods=["POST"])
def send_email(collection_id):
    collections = load_collections()
    collection = collections.get(collection_id)
    if not collection:
        flash("Coleta não encontrada.", "danger")
        return redirect(url_for("index"))

    email = request.form.get("email", "").strip()
    if not email:
        flash("Informe o email de destino.", "warning")
        return redirect(url_for("report", collection_id=collection_id))

    report_path = create_pdf_report(collection)
    subject = f"Relatório de coleta {collection['equipe']} - {collection['id']}"
    body = (
        f"Relatório da coleta {collection['equipe']}\n"
        f"Operador: {collection['operador']}\n"
        f"Início: {collection['start']}\n"
        f"Término: {collection.get('end', 'em andamento')}\n"
        f"Total de registros: {len(collection['items'])}\n\n"
        f"O arquivo está em anexo."
    )

    try:
        send_email_report(email, subject, body, report_path)
        flash(f"Email enviado para {email}.", "success")
    except Exception as exc:
        flash(f"Erro ao enviar email: {exc}", "danger")
    return redirect(url_for("report", collection_id=collection_id))


@app.route("/start", methods=["POST"])
def start_collection():
    equipe = request.form.get("equipe", "").strip()
    operador = request.form.get("operador", "").strip()
    if not equipe or not operador:
        flash("Informe equipe e operador para iniciar a coleta.", "warning")
        return redirect(url_for("index"))

    collection_id = str(uuid.uuid4())
    now = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
    collection = {
        "id": collection_id,
        "equipe": equipe,
        "operador": operador,
        "start": now,
        "updated": now,
        "items": [],
        "status": "aberta",
    }
    collections = load_collections()
    collections[collection_id] = collection
    save_collections(collections)
    flash("Coleta iniciada. Escaneie medidores pelo celular ou pelo PC.", "success")
    return redirect(url_for("collect", collection_id=collection_id))


@app.route("/collect/<collection_id>", methods=["GET", "POST"])
def collect(collection_id):
    collections = load_collections()
    collection = collections.get(collection_id)
    if not collection:
        flash("Coleta não encontrada.", "danger")
        return redirect(url_for("index"))

    if request.method == "POST":
        medidor = request.form.get("medidor", "").strip()
        tipo = request.form.get("tipo", "").strip()
        status = request.form.get("status", "").strip()
        observacoes = request.form.get("observacoes", "").strip()
        if not medidor or not tipo or not status:
            flash("Preencha medidor, tipo e status antes de enviar.", "warning")
            return redirect(url_for("collect", collection_id=collection_id))

        try:
            create_entry(collection, medidor, tipo, status, observacoes)
            collections[collection_id] = collection
            save_collections(collections)
            flash("Registro salvo na planilha mestre e na planilha da equipe.", "success")
        except Exception as exc:
            flash(f"Erro ao salvar o registro: {exc}", "danger")
        return redirect(url_for("collect", collection_id=collection_id))

    last_items = list(reversed(collection["items"]))[:10]
    return render_template(
        "collector.html",
        collection=collection,
        last_items=last_items,
        reportlab_available=REPORTLAB_AVAILABLE,
    )


@app.route("/finalize/<collection_id>", methods=["POST"])
def finalize(collection_id):
    collections = load_collections()
    collection = collections.get(collection_id)
    if not collection:
        flash("Coleta não encontrada.", "danger")
        return redirect(url_for("index"))
    if collection.get("status") == "finalizada":
        flash("Coleta já foi finalizada.", "info")
        return redirect(url_for("report", collection_id=collection_id))

    collection["end"] = datetime.datetime.now().isoformat(sep=" ", timespec="seconds")
    collection["status"] = "finalizada"
    save_collections(collections)
    report_path = create_pdf_report(collection)
    flash(f"Coleta finalizada. Relatório gerado: {report_path.name}", "success")
    return redirect(url_for("report", collection_id=collection_id))


@app.route("/report/<collection_id>")
def report(collection_id):
    collections = load_collections()
    collection = collections.get(collection_id)
    if not collection:
        flash("Coleta não encontrada.", "danger")
        return redirect(url_for("index"))

    filename = f"relatorio_{collection_id}.pdf"
    if not REPORTLAB_AVAILABLE:
        filename = f"relatorio_{collection_id}.txt"

    report_path = REPORTS_DIR / filename
    if not report_path.exists():
        report_path = create_pdf_report(collection)

    email_config = load_email_config()
    share_url = request.host_url.rstrip("/") + url_for("download", filename=report_path.name)
    return render_template(
        "report.html",
        collection=collection,
        report_path=report_path.name,
        reportlab_available=REPORTLAB_AVAILABLE,
        email_enabled=is_email_configured(email_config),
        share_url=share_url,
    )


@app.route("/download/<filename>")
def download(filename):
    file_path = REPORTS_DIR / filename
    if not file_path.exists():
        flash("Relatório não encontrado.", "danger")
        return redirect(url_for("index"))
    return send_file(file_path, as_attachment=True)


@app.route("/history")
def history():
    collections = load_collections()
    ordered = sorted(collections.values(), key=lambda c: c["start"], reverse=True)
    return render_template("history.html", collections=ordered)


def ensure_data_files() -> None:
    ensure_workbook(MASTER_FILE)
    if not COLLECTIONS_FILE.exists():
        save_collections({})
    ensure_email_config()
    ensure_server_config()


if __name__ == "__main__":
    ensure_data_files()
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "False").lower() in ("1", "true", "yes")
    app.run(host="0.0.0.0", port=port, debug=debug)
