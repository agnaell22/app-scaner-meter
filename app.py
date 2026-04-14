import os
import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook

MASTER_FILE = "master.xlsx"
USER_FILE_TEMPLATE = "planilha_{equipe}.xlsx"
HEADERS = ["Data", "Equipe", "Medidor", "Tipo", "Status", "Observações"]


def ensure_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        wb = load_workbook(path)
        if "Dados" not in wb.sheetnames:
            ws = wb.create_sheet("Dados")
            ws.append(HEADERS)
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.append(HEADERS)
    wb.save(path)
    return wb


def append_row(path: str, row: list[str]) -> None:
    wb = ensure_workbook(path)
    ws = wb["Dados"]
    ws.append(row)
    wb.save(path)


def save_entry():
    equipe = entry_equipe.get().strip()
    medidor = entry_medidor.get().strip()
    tipo = combo_tipo.get().strip()
    status = combo_status.get().strip()
    observacoes = text_observacoes.get("1.0", tk.END).strip()

    if not equipe:
        messagebox.showwarning("Dados incompletos", "Informe a equipe antes de salvar.")
        return
    if not medidor:
        messagebox.showwarning("Dados incompletos", "Informe o medidor ou faça a leitura do código de barras.")
        return
    if not tipo:
        messagebox.showwarning("Dados incompletos", "Informe o tipo do medidor.")
        return
    if not status:
        messagebox.showwarning("Dados incompletos", "Informe o status do medidor.")
        return

    data = entry_data.get().strip() or datetime.date.today().isoformat()
    row = [data, equipe, medidor, tipo, status, observacoes]

    try:
        append_row(MASTER_FILE, row)
        user_file = USER_FILE_TEMPLATE.format(equipe=equipe.replace(" ", "_"))
        append_row(user_file, row)
        label_status.config(text=f"Registro salvo em {MASTER_FILE} e {user_file}.", foreground="green")
        clear_fields()
    except Exception as exc:
        messagebox.showerror("Erro", f"Não foi possível salvar os dados: {exc}")


def clear_fields():
    entry_medidor.delete(0, tk.END)
    combo_tipo.set("")
    combo_status.set("")
    text_observacoes.delete("1.0", tk.END)
    entry_medidor.focus()


def on_enter_pressed(event):
    save_entry()


root = tk.Tk()
root.title("Coleta de Medidores")
root.geometry("520x440")
root.resizable(False, False)

frame = ttk.Frame(root, padding=16)
frame.pack(fill=tk.BOTH, expand=True)

label_data = ttk.Label(frame, text="Data:")
label_data.grid(row=0, column=0, sticky=tk.W, pady=4)
entry_data = ttk.Entry(frame)
entry_data.insert(0, datetime.date.today().isoformat())
entry_data.grid(row=0, column=1, columnspan=2, sticky=tk.EW, pady=4)

label_equipe = ttk.Label(frame, text="Equipe:")
label_equipe.grid(row=1, column=0, sticky=tk.W, pady=4)
entry_equipe = ttk.Entry(frame)
entry_equipe.grid(row=1, column=1, columnspan=2, sticky=tk.EW, pady=4)

label_medidor = ttk.Label(frame, text="Medidor / Código de barras:")
label_medidor.grid(row=2, column=0, sticky=tk.W, pady=4)
entry_medidor = ttk.Entry(frame)
entry_medidor.grid(row=2, column=1, columnspan=2, sticky=tk.EW, pady=4)
entry_medidor.bind("<Return>", on_enter_pressed)

label_tipo = ttk.Label(frame, text="Tipo do medidor:")
label_tipo.grid(row=3, column=0, sticky=tk.W, pady=4)
combo_tipo = ttk.Combobox(frame, values=["Elétrico", "Hídrico", "Gás", "Outros"], state="readonly")
combo_tipo.grid(row=3, column=1, columnspan=2, sticky=tk.EW, pady=4)

label_status = ttk.Label(frame, text="Status:")
label_status.grid(row=4, column=0, sticky=tk.W, pady=4)
combo_status = ttk.Combobox(frame, values=["Em uso", "Retirado", "Devolvido", "Aguardando baixa"], state="readonly")
combo_status.grid(row=4, column=1, columnspan=2, sticky=tk.EW, pady=4)

label_observacoes = ttk.Label(frame, text="Observações / Controle:")
label_observacoes.grid(row=5, column=0, sticky=tk.NW, pady=4)
text_observacoes = tk.Text(frame, height=6, width=44)
text_observacoes.grid(row=5, column=1, columnspan=2, sticky=tk.EW, pady=4)

button_save = ttk.Button(frame, text="Registrar entrada", command=save_entry)
button_save.grid(row=6, column=1, sticky=tk.EW, pady=12)

button_clear = ttk.Button(frame, text="Limpar campos", command=clear_fields)
button_clear.grid(row=6, column=2, sticky=tk.EW, pady=12)

label_status = ttk.Label(frame, text="Preencha os campos e pressione Registrar entrada.")
label_status.grid(row=7, column=0, columnspan=3, sticky=tk.W, pady=8)

for col in range(3):
    frame.columnconfigure(col, weight=1)

# Garantir arquivos iniciais
ensure_workbook(MASTER_FILE)
root.mainloop()
